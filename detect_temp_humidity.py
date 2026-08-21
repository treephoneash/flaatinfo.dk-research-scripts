from openpyxl import load_workbook
import requests
from datetime import datetime, timedelta
import sys

file_path = "input.xlsx"
log_file_path = "error_log.txt"

wb = load_workbook(file_path)
ws = wb.active

start_row = 7001
end_row = 9000
total = end_row - start_row + 1

cache = {}  # cache API responses keyed by (lat, lon, date)

with open(log_file_path, "w") as log_file:

    for idx, row in enumerate(range(start_row, end_row + 1), start=1):

        sys.stdout.write(f"\rProcessing row {idx} of {total}")
        sys.stdout.flush()

        try:
            date_str = str(ws[f"C{row}"].value)
            lat = ws[f"N{row}"].value
            lon = ws[f"O{row}"].value

            if not date_str or lat is None or lon is None:
                raise ValueError("Missing date, latitude, or longitude")

            dt = datetime.strptime(date_str, "%Y-%m-%d %H:%M")

            # Custom rounding rule
            if 1 <= dt.minute <= 30:
                rounded = dt.replace(minute=0, second=0, microsecond=0)
            elif 31 <= dt.minute <= 59:
                rounded = (dt + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
            else:  # exactly minute 00
                rounded = dt.replace(second=0, microsecond=0)

            date_only = rounded.strftime("%Y-%m-%d")
            hour_target = rounded.strftime("%Y-%m-%dT%H:00")

            cache_key = f"{lat}_{lon}_{date_only}"

            if cache_key not in cache:

                url = (
                    "https://archive-api.open-meteo.com/v1/archive"
                    f"?latitude={lat}&longitude={lon}"
                    f"&start_date={date_only}&end_date={date_only}"
                    "&hourly=temperature_2m,relative_humidity_2m"
                )

                response = requests.get(url, timeout=20)
                response.raise_for_status()
                data = response.json()

                times = data["hourly"]["time"]
                temps = data["hourly"]["temperature_2m"]
                humidity = data["hourly"]["relative_humidity_2m"]

                # convert to fast lookup dictionary
                hourly_map = {
                    t: (temp, hum)
                    for t, temp, hum in zip(times, temps, humidity)
                }

                cache[cache_key] = hourly_map

            hourly_map = cache[cache_key]

            if hour_target in hourly_map:
                temp, hum = hourly_map[hour_target]
                ws[f"R{row}"] = temp
                ws[f"S{row}"] = hum
            else:
                raise ValueError(f"No data for hour {hour_target}")

        except Exception as e:
            ws[f"R{row}"] = None
            ws[f"S{row}"] = None
            log_file.write(
                f"Row {row} error: {e} | Date: {date_str} | Lat: {lat} | Lon: {lon}\n"
            )

wb.save(file_path)

print("\nFinished processing. Check 'error_log.txt' for failed rows.")
